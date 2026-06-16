#!/usr/bin/env python3
"""
Institution Paper Tracker — XLSX Generator
Generates xlsx matching the standard 两院署名论文 template format.

Usage: python3 gen_xlsx.py <output_path>

Expected input: a JSON file at /tmp/paper_results.json with structure:
{
  "papers": [
    {
      "person_type": "学生" or "导师",
      "name": "Author Name",
      "institution": "原高校/机构",
      "title": "Paper Title",
      "date": "YYYY-MM-DD",
      "arxiv_id": "2603.20611" or "",
      "arxiv_url": "https://arxiv.org/abs/2603.20611" or "",
      "authors": "Author1; Author2; Author3"
    },
    ...
  ]
}
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: python3 -m pip install --user openpyxl")
    sys.exit(1)

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP = Alignment(vertical="center", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["人员类型", "姓名", "原高校/机构", "论文标题", "提交日期",
           "arXiv ID", "arXiv链接", "作者列表"]
COL_WIDTHS = [8, 10, 16, 55, 12, 12, 35, 55]


def build_xlsx(papers, output_path, sheet_name="两院署名论文"):
    # Sort: 学生 first, then 导师; within each group, date descending
    type_order = {"学生": 0, "导师": 1}
    papers.sort(key=lambda p: (type_order.get(p["person_type"], 9), p["date"]), reverse=False)
    # Reverse date within same type
    papers_sorted = []
    for ptype in ["学生", "导师"]:
        group = [p for p in papers if p["person_type"] == ptype]
        group.sort(key=lambda p: p["date"], reverse=True)
        papers_sorted.extend(group)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, paper in enumerate(papers_sorted, 2):
        values = [
            paper["person_type"],
            paper["name"],
            paper["institution"],
            paper["title"],
            paper["date"],
            paper.get("arxiv_id", ""),
            paper.get("arxiv_url", ""),
            paper["authors"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-fit row heights (approximate)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")
    print(f"   Sheet: {sheet_name} | Rows: {len(papers_sorted)}")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser("~/Desktop/papers.xlsx")
    data_path = "/tmp/paper_results.json"

    if not os.path.exists(data_path):
        print(f"ERROR: No data file at {data_path}")
        sys.exit(1)

    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    build_xlsx(data["papers"], output_path)