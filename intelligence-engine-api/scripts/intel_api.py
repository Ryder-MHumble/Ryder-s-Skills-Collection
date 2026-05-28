#!/usr/bin/env python3
"""Small helper for the internal Intelligence Engine API.

Default base URL: http://10.1.132.21:8001
Override with INTEL_ENGINE_BASE_URL or --base-url.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

DEFAULT_BASE_URL = os.environ.get("INTEL_ENGINE_BASE_URL", "http://10.1.132.21:8001")

ALIASES = {
    "openapi": "/openapi.json",
    "dimensions": "/api/dimensions",
    "articles": "/api/articles",
    "article-stats": "/api/articles/stats",
    "sources": "/api/sources",
    "sources-catalog": "/api/sources/catalog",
    "sources-facets": "/api/sources/facets",
    "sources-items": "/api/sources/items",
    "social-posts": "/api/social-posts",
    "wechat-mp-posts": "/api/social-posts/wechat-mp",
    "social-stats": "/api/social-posts/stats",
    "daily-metrics": "/api/intel/daily-briefing/metrics",
    "daily-today": "/api/intel/daily-briefing/today",
    "daily-report": "/api/intel/daily-briefing/report",
    "policy-feed": "/api/intel/policy/feed",
    "policy-stats": "/api/intel/policy/stats",
    "policy-opportunities": "/api/intel/policy/opportunities",
    "university-overview": "/api/intel/university/overview",
    "university-feed": "/api/intel/university/feed",
    "university-sources": "/api/intel/university/sources",
    "tech-topics": "/api/intel/tech-frontier/topics",
    "tech-signals": "/api/intel/tech-frontier/signals",
    "tech-stats": "/api/intel/tech-frontier/stats",
    "personnel-feed": "/api/intel/personnel/feed",
    "personnel-stats": "/api/intel/personnel/stats",
    "sentiment-overview": "/api/sentiment/overview",
    "sentiment-feed": "/api/sentiment/feed",
    "papers": "/api/papers",
    "paper-sources": "/api/papers/sources",
    "venues": "/api/venues",
    "venues-stats": "/api/venues/stats",
    "scholars": "/api/scholars",
    "scholars-stats": "/api/scholars/stats",
    "institutions": "/api/institutions",
    "institutions-stats": "/api/institutions/stats",
    "health": "/api/health",
    "crawl-status": "/api/health/crawl-status",
    "pipeline-status": "/api/health/pipeline-status",
}


def parse_params(pairs: list[str]) -> dict[str, str]:
    params: dict[str, str] = {}
    for pair in pairs:
        if "=" not in pair:
            raise SystemExit(f"--param must be key=value, got: {pair}")
        key, value = pair.split("=", 1)
        if not key:
            raise SystemExit(f"empty parameter name in: {pair}")
        params[key] = value
    return params


def resolve_path(alias_or_path: str) -> str:
    if alias_or_path in ALIASES:
        return ALIASES[alias_or_path]
    if alias_or_path.startswith("/"):
        return alias_or_path
    raise SystemExit(f"unknown alias/path: {alias_or_path}. Use --list-aliases to inspect aliases.")


def api_get(base_url: str, alias_or_path: str, params: dict[str, Any] | None = None, timeout: int = 30) -> Any:
    path = resolve_path(alias_or_path)
    base = base_url.rstrip("/") + "/"
    url = urljoin(base, path.lstrip("/"))
    if params:
        clean_params = {k: v for k, v in params.items() if v is not None}
        url += "?" + urlencode(clean_params, doseq=True)
    req = Request(url, headers={"Accept": "application/json", "User-Agent": "codex-intel-engine-skill/1.0"})
    try:
        with urlopen(req, timeout=timeout) as resp:
            body = resp.read().decode("utf-8-sig")
            if not body:
                return None
            return json.loads(body)
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:1000]
        raise SystemExit(f"HTTP {exc.code} for {url}\n{detail}") from exc
    except URLError as exc:
        raise SystemExit(f"Network error for {url}: {exc}") from exc


def write_json(data: Any, out: str | None, pretty: bool) -> None:
    text = json.dumps(data, ensure_ascii=False, indent=2 if pretty else None)
    if out:
        with open(out, "w", encoding="utf-8") as f:
            f.write(text + "\n")
    else:
        print(text)


def short_text(value: Any, limit: int = 90) -> str:
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else text[: limit - 1] + "?"


def first_items(data: Any, n: int = 3) -> list[dict[str, Any]]:
    if isinstance(data, dict):
        items = data.get("items") or []
    elif isinstance(data, list):
        items = data
    else:
        items = []
    return items[:n]


def poster_snapshot(base_url: str, target_date: str, days: int, timeout: int) -> dict[str, Any]:
    end = datetime.strptime(target_date, "%Y-%m-%d").date()
    start = end - timedelta(days=max(days, 1) - 1)
    date_from = start.isoformat()
    date_to = end.isoformat()

    result: dict[str, Any] = {
        "date": target_date,
        "date_window": {"date_from": date_from, "date_to": date_to, "days": days},
        "base_url": base_url.rstrip("/"),
        "endpoints_used": [],
    }

    def get(alias: str, params: dict[str, Any] | None = None) -> Any:
        result["endpoints_used"].append({"alias": alias, "path": resolve_path(alias), "params": params or {}})
        return api_get(base_url, alias, params=params, timeout=timeout)

    metrics = get("daily-metrics", {"target_date": target_date})
    result["daily_metrics"] = metrics

    sources = get("sources-catalog", {"page_size": 1, "include_facets": "true"})
    result["source_coverage"] = {
        "total_sources": sources.get("total_sources") if isinstance(sources, dict) else None,
        "filtered_sources": sources.get("filtered_sources") if isinstance(sources, dict) else None,
        "facets": sources.get("facets") if isinstance(sources, dict) else None,
    }

    social_stats = get("social-stats", {"group_by": "platform"})
    result["social_stats"] = social_stats

    paper_list = get("papers", {"page_size": 5, "sort_by": "publication_date", "order": "desc"})
    result["paper_summary"] = {
        "total": paper_list.get("total") if isinstance(paper_list, dict) else None,
        "recent": [
            {
                "date": (item.get("publication_date") or "")[:10],
                "source": ((item.get("source") or {}).get("name")),
                "title": short_text(item.get("title"), 110),
            }
            for item in first_items(paper_list, 5)
        ],
    }

    policy = get("policy-feed", {"limit": 5})
    result["policy_cards"] = [
        {
            "date": item.get("date"),
            "category": item.get("category"),
            "importance": item.get("importance"),
            "source": item.get("source") or item.get("source_name"),
            "title": short_text(item.get("title"), 100),
            "summary": short_text(item.get("summary"), 140),
        }
        for item in first_items(policy, 5)
    ]

    tech = get("tech-signals", {"limit": 6})
    result["tech_signal_cards"] = []
    for item in first_items(tech, 6):
        data = item.get("data", {}) if isinstance(item, dict) else {}
        result["tech_signal_cards"].append(
            {
                "date": item.get("date"),
                "topic": item.get("parentTopicName"),
                "kind": item.get("kind"),
                "source": data.get("source") or data.get("author") or data.get("source_name"),
                "title": short_text(data.get("title") or data.get("content") or data.get("text"), 110),
            }
        )

    result["article_samples"] = {}
    for dim in ["beijing_policy", "industry", "technology", "universities", "talent", "events"]:
        articles = get(
            "articles",
            {
                "dimension": dim,
                "date_from": date_from,
                "date_to": date_to,
                "sort_by": "published_at",
                "order": "desc",
                "page_size": 3,
            },
        )
        result["article_samples"][dim] = {
            "total": articles.get("total") if isinstance(articles, dict) else None,
            "items": [
                {
                    "date": (item.get("published_at") or item.get("crawled_at") or "")[:10],
                    "source": item.get("source_name") or item.get("source_id"),
                    "title": short_text(item.get("title"), 110),
                }
                for item in first_items(articles, 3)
            ],
        }

    social = get("social-posts", {"page_size": 5, "sort_by": "published_at", "order": "desc"})
    result["social_recent"] = [
        {
            "date": (item.get("published_at") or "")[:10],
            "platform": item.get("platform"),
            "author": item.get("author_username") or item.get("author_display_name"),
            "text": short_text(item.get("content_text"), 140),
        }
        for item in first_items(social, 5)
    ]

    return result


def dedupe_strings(values: list[Any]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        text = str(value or "").strip()
        if text and text not in seen:
            seen.add(text)
            out.append(text)
    return out


def affiliation_summary(affiliations: list[Any]) -> str:
    values: list[str] = []
    for item in affiliations or []:
        if isinstance(item, dict):
            name = item.get("author_name") or ""
            affiliation = item.get("affiliation") or ""
            values.append(f"{name}: {affiliation}" if name and affiliation else affiliation or name)
        else:
            values.append(str(item))
    return "; ".join(dedupe_strings(values))


def export_papers_xlsx(
    base_url: str,
    out: str,
    timeout: int,
    venue: str | None = None,
    venue_year: str | None = None,
    q: str | None = None,
    source_name: str | None = None,
    source_id: str | None = None,
    affiliation: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    has_abstract: str | None = None,
    sort_by: str = "publication_date",
    order: str = "desc",
) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:  # pragma: no cover - environment dependent
        raise SystemExit(f"openpyxl is required for xlsx export: {exc}") from exc

    base_params = {
        "venue": venue,
        "venue_year": venue_year,
        "q": q,
        "source_name": source_name,
        "source_id": source_id,
        "affiliation": affiliation,
        "date_from": date_from,
        "date_to": date_to,
        "has_abstract": has_abstract,
        "sort_by": sort_by,
        "order": order,
        "page_size": 100,
    }
    first = api_get(base_url, "papers", {**base_params, "page": 1}, timeout=timeout)
    total = int(first.get("total") or 0) if isinstance(first, dict) else 0
    pages = (total + 99) // 100 if total else 1
    items = list(first.get("items") or []) if isinstance(first, dict) else []
    for page in range(2, pages + 1):
        data = api_get(base_url, "papers", {**base_params, "page": page}, timeout=timeout)
        items.extend(data.get("items") or [])
        if page % 10 == 0 or page == pages:
            print(f"fetched page {page}/{pages}", file=sys.stderr)

    seen: set[str] = set()
    unique_items: list[dict[str, Any]] = []
    for item in items:
        key = item.get("paper_id") or item.get("canonical_uid") or item.get("detail_url") or item.get("title")
        if key in seen:
            continue
        seen.add(key)
        unique_items.append(item)

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    headers = [
        "paper_id",
        "title",
        "authors",
        "affiliations",
        "venue",
        "venue_year",
        "track",
        "publication_date",
        "doi",
        "abstract",
        "detail_url",
        "pdf_url",
        "source_name",
        "ingested_at",
        "updated_at",
    ]
    ws.append(headers)
    for item in unique_items:
        source = item.get("source") or {}
        ws.append(
            [
                item.get("paper_id"),
                item.get("title"),
                "; ".join(dedupe_strings(item.get("authors") or [])),
                affiliation_summary(item.get("affiliations") or []),
                item.get("venue"),
                item.get("venue_year"),
                item.get("track"),
                item.get("publication_date"),
                item.get("doi"),
                item.get("abstract"),
                item.get("detail_url"),
                item.get("pdf_url"),
                source.get("name"),
                item.get("ingested_at"),
                item.get("updated_at"),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 34,
        "B": 70,
        "C": 48,
        "D": 70,
        "E": 12,
        "F": 12,
        "G": 45,
        "H": 25,
        "I": 22,
        "J": 90,
        "K": 55,
        "L": 55,
        "M": 20,
        "N": 25,
        "O": 25,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    meta = wb.create_sheet("Metadata")
    metadata = [
        ["exported_at", datetime.now().isoformat(timespec="seconds")],
        ["api_base_url", base_url.rstrip("/")],
        ["endpoint", "/api/papers"],
        ["query_params", json.dumps({k: v for k, v in base_params.items() if v is not None}, ensure_ascii=False)],
        ["api_total", total],
        ["exported_rows", len(unique_items)],
        ["pages", pages],
    ]
    for row in metadata:
        meta.append(row)
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 120

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"output": str(out_path), "api_total": total, "exported_rows": len(unique_items), "pages": pages}


def values_at_path(obj: Any, path: str) -> list[Any]:
    """Return one or more values for a dotted path; lists are expanded."""
    if path in ("", "."):
        return [obj]
    values = [obj]
    for part in path.split("."):
        if part.endswith("[]"):
            part = part[:-2]
        next_values: list[Any] = []
        for value in values:
            if isinstance(value, dict):
                child = value.get(part)
                if isinstance(child, list):
                    next_values.extend(child)
                elif child is not None:
                    next_values.append(child)
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        child = item.get(part)
                        if isinstance(child, list):
                            next_values.extend(child)
                        elif child is not None:
                            next_values.append(child)
                    elif part == "*":
                        next_values.append(item)
        values = next_values
        if not values:
            break
    return values


def value_at_path(obj: Any, path: str) -> Any:
    values = values_at_path(obj, path)
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    return "; ".join(dedupe_strings(values))


def flatten_record(obj: Any, prefix: str = "", out: dict[str, Any] | None = None) -> dict[str, Any]:
    out = out or {}
    if isinstance(obj, dict):
        for key, value in obj.items():
            child_key = f"{prefix}.{key}" if prefix else str(key)
            flatten_record(value, child_key, out)
    elif isinstance(obj, list):
        if all(not isinstance(item, (dict, list)) for item in obj):
            out[prefix] = "; ".join(dedupe_strings(obj))
        else:
            out[prefix] = json.dumps(obj, ensure_ascii=False)
    else:
        out[prefix] = obj
    return out


def parse_fields(fields: str | None) -> list[tuple[str, str]]:
    if not fields:
        return []
    parsed: list[tuple[str, str]] = []
    for raw in fields.split(","):
        part = raw.strip()
        if not part:
            continue
        if "=" in part:
            label, path = part.split("=", 1)
            parsed.append((label.strip(), path.strip()))
        else:
            parsed.append((part, part))
    return parsed


def cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def extract_items(data: Any, items_path: str = "items") -> list[Any]:
    if isinstance(data, list):
        return data
    if not isinstance(data, dict):
        return []
    if items_path:
        found = value_at_path(data, items_path)
        if isinstance(found, list):
            return found
    items = data.get("items")
    return items if isinstance(items, list) else []


def fetch_collection(
    base_url: str,
    endpoint: str,
    params: dict[str, Any],
    timeout: int,
    pagination: str = "none",
    page_size: int = 100,
    max_pages: int | None = None,
    items_path: str = "items",
) -> tuple[list[Any], dict[str, Any]]:
    """Fetch an endpoint and optionally walk page/offset pagination."""
    if pagination == "none":
        data = api_get(base_url, endpoint, params=params, timeout=timeout)
        items = extract_items(data, items_path)
        meta: dict[str, Any] = {"pages_fetched": 1, "items_fetched": len(items)}
        if isinstance(data, dict):
            meta.update({
                "total": data.get("total"),
                "page": data.get("page"),
                "page_size": data.get("page_size"),
                "total_pages": data.get("total_pages"),
                "response_keys": list(data.keys()),
            })
        elif isinstance(data, list):
            meta["root_list_count"] = len(data)
        return items, meta

    clean_page_size = min(max(page_size, 1), 100)
    items: list[Any] = []
    pages_fetched = 0
    total = None
    total_pages = None
    next_offset = 0
    page = 1
    while True:
        request_params = dict(params)
        if pagination == "page":
            request_params.update({"page": page, "page_size": clean_page_size})
        elif pagination == "offset":
            request_params.update({"offset": next_offset, "limit": clean_page_size})
        else:
            raise SystemExit("--pagination must be none, page, or offset")
        data = api_get(base_url, endpoint, params=request_params, timeout=timeout)
        batch = extract_items(data, items_path)
        items.extend(batch)
        pages_fetched += 1
        if isinstance(data, dict):
            total = data.get("total", total)
            total_pages = data.get("total_pages", total_pages)
            if pagination == "offset":
                if data.get("has_more") is False:
                    break
                next_offset = int(data.get("next_offset") or (next_offset + clean_page_size))
            else:
                if total_pages and page >= int(total_pages):
                    break
        if max_pages and pages_fetched >= max_pages:
            break
        if pagination == "page":
            if not batch or len(batch) < clean_page_size:
                break
            page += 1
        elif pagination == "offset":
            if not batch or len(batch) < clean_page_size:
                break
    return items, {"total": total, "total_pages": total_pages, "pages_fetched": pages_fetched}


def match_where(item: Any, where_expr: str) -> bool:
    operators = ["!=", ">=", "<=", "~", "=", ">", "<"]
    op = next((candidate for candidate in operators if candidate in where_expr), None)
    if not op:
        raise SystemExit(f"where expression must contain one of {operators}: {where_expr}")
    left, right = where_expr.split(op, 1)
    actual = value_at_path(item, left.strip())
    expected = right.strip()
    actual_text = "" if actual is None else str(actual)
    if op == "~":
        return expected.lower() in actual_text.lower()
    if op == "=":
        return actual_text == expected
    if op == "!=":
        return actual_text != expected
    try:
        actual_num = float(actual_text)
        expected_num = float(expected)
    except ValueError:
        return False
    if op == ">":
        return actual_num > expected_num
    if op == "<":
        return actual_num < expected_num
    if op == ">=":
        return actual_num >= expected_num
    if op == "<=":
        return actual_num <= expected_num
    return False


def apply_local_ops(
    items: list[Any],
    where: list[str] | None = None,
    sort: str | None = None,
    sort_desc: bool = False,
    limit: int | None = None,
) -> list[Any]:
    result = list(items)
    for expr in where or []:
        result = [item for item in result if match_where(item, expr)]
    if sort:
        def sort_key(item: Any) -> tuple[int, Any]:
            value = value_at_path(item, sort)
            if value is None:
                return (1, "")
            try:
                return (0, float(value))
            except (TypeError, ValueError):
                return (0, str(value))
        result.sort(key=sort_key, reverse=sort_desc)
    if limit is not None:
        result = result[:limit]
    return result


def rows_from_items(items: list[Any], fields: str | None = None) -> tuple[list[str], list[list[Any]]]:
    parsed_fields = parse_fields(fields)
    if not parsed_fields:
        keys: list[str] = []
        for item in items[:100]:
            for key in flatten_record(item).keys():
                if key not in keys:
                    keys.append(key)
        parsed_fields = [(key, key) for key in keys]
    headers = [label for label, _ in parsed_fields]
    rows = [[cell_value(value_at_path(item, path)) for _, path in parsed_fields] for item in items]
    return headers, rows


def write_rows_xlsx(out: str, headers: list[str], rows: list[list[Any]], metadata: list[list[Any]] | None = None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:  # pragma: no cover - environment dependent
        raise SystemExit(f"openpyxl is required for xlsx export: {exc}") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col_cells[:100])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if metadata:
        meta = wb.create_sheet("Metadata")
        for row in metadata:
            meta.append(row)
        meta.column_dimensions["A"].width = 24
        meta.column_dimensions["B"].width = 120
    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def generic_export_xlsx(
    base_url: str,
    endpoint: str,
    params: dict[str, Any],
    out: str,
    timeout: int,
    items_path: str = "items",
    pagination: str = "none",
    page_size: int = 100,
    max_pages: int | None = None,
    fields: str | None = None,
    where: list[str] | None = None,
    sort: str | None = None,
    sort_desc: bool = False,
    limit: int | None = None,
) -> dict[str, Any]:
    items, fetch_meta = fetch_collection(base_url, endpoint, params, timeout, pagination, page_size, max_pages, items_path)
    items = apply_local_ops(items, where=where, sort=sort, sort_desc=sort_desc, limit=limit)
    headers, rows = rows_from_items(items, fields)
    metadata = [
        ["exported_at", datetime.now().isoformat(timespec="seconds")],
        ["api_base_url", base_url.rstrip("/")],
        ["endpoint", endpoint],
        ["path", resolve_path(endpoint)],
        ["query_params", json.dumps(params, ensure_ascii=False)],
        ["pagination", pagination],
        ["fetch_meta", json.dumps(fetch_meta, ensure_ascii=False)],
        ["exported_rows", len(rows)],
        ["fields", fields or "(auto)"],
        ["where", "; ".join(where or [])],
        ["sort", sort or ""],
    ]
    write_rows_xlsx(out, headers, rows, metadata)
    return {"output": out, "exported_rows": len(rows), "fetch_meta": fetch_meta}


def write_rows_json(out: str | None, headers: list[str], rows: list[list[Any]]) -> None:
    data = [dict(zip(headers, row)) for row in rows]
    write_json(data, out, pretty=True)


def generic_chart(
    base_url: str,
    endpoint: str,
    params: dict[str, Any],
    out: str,
    timeout: int,
    x_field: str,
    y_field: str | None = None,
    chart_type: str = "bar",
    aggregate: str = "count",
    top: int | None = 20,
    items_path: str = "items",
    pagination: str = "none",
    page_size: int = 100,
    max_pages: int | None = None,
) -> dict[str, Any]:
    try:
        import matplotlib.pyplot as plt
    except Exception as exc:  # pragma: no cover - environment dependent
        raise SystemExit(f"matplotlib is required for chart export: {exc}") from exc

    items, fetch_meta = fetch_collection(base_url, endpoint, params, timeout, pagination, page_size, max_pages, items_path)
    buckets: dict[str, float] = {}
    for item in items:
        keys = values_at_path(item, x_field)
        if not keys:
            keys = ["(empty)"]
        for key in keys:
            label = str(key)
            if y_field:
                raw_value = value_at_path(item, y_field)
                try:
                    value = float(raw_value or 0)
                except (TypeError, ValueError):
                    value = 0.0
            else:
                value = 1.0
            if aggregate == "count":
                buckets[label] = buckets.get(label, 0) + 1
            elif aggregate == "sum":
                buckets[label] = buckets.get(label, 0) + value
            else:
                raise SystemExit("--aggregate must be count or sum")
    pairs = sorted(buckets.items(), key=lambda kv: kv[1], reverse=True)
    if top:
        pairs = pairs[:top]
    labels = [p[0] for p in pairs]
    values = [p[1] for p in pairs]
    plt.figure(figsize=(12, 7))
    if chart_type == "bar":
        plt.bar(labels, values, color="#2f6f6d")
        plt.xticks(rotation=35, ha="right")
    elif chart_type == "line":
        plt.plot(labels, values, marker="o", color="#2f6f6d")
        plt.xticks(rotation=35, ha="right")
    elif chart_type == "pie":
        plt.pie(values, labels=labels, autopct="%1.1f%%")
    else:
        raise SystemExit("--chart-type must be bar, line, or pie")
    plt.title(f"{endpoint}: {aggregate} by {x_field}")
    plt.tight_layout()
    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_path, dpi=160)
    plt.close()
    return {"output": str(out_path), "points": len(pairs), "items": len(items), "fetch_meta": fetch_meta}


def parse_cn_date(ts: str | None) -> str | None:
    if not ts:
        return None
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except ValueError:
        return None
    return dt.astimezone(timezone(timedelta(hours=8))).date().isoformat()


def rank_social_posts(
    base_url: str,
    params: dict[str, Any],
    timeout: int,
    date_filter: str | None = None,
    limit: int = 50,
    out_xlsx: str | None = None,
    out_json: str | None = None,
    max_pages: int = 5,
) -> dict[str, Any]:
    items, fetch_meta = fetch_collection(
        base_url,
        "social-posts",
        {"platform": "x", "sort_by": "published_at", "order": "desc", **params},
        timeout,
        pagination="page",
        page_size=100,
        max_pages=max_pages,
    )
    if date_filter:
        items = [item for item in items if parse_cn_date(item.get("published_at")) == date_filter]
    ranked = []
    for item in items:
        likes = int(item.get("like_count") or 0)
        replies = int(item.get("reply_count") or 0)
        reposts = int(item.get("repost_count") or 0)
        quotes = int(item.get("quote_count") or 0)
        views = int(item.get("view_count") or 0)
        bookmarks = int(item.get("bookmark_count") or 0)
        score = likes + replies * 1.5 + reposts * 2 + quotes * 2 + bookmarks * 1.5 + views * 0.01
        ranked.append({**item, "heat_score": round(score, 2)})
    ranked.sort(key=lambda item: item["heat_score"], reverse=True)
    ranked = ranked[:limit]
    fields = (
        "heat_score,published_at,author=author_username,name=author_display_name,post_type,"
        "content=content_text,like_count,reply_count,repost_count,quote_count,view_count,bookmark_count,post_url"
    )
    headers, rows = rows_from_items(ranked, fields)
    if out_xlsx:
        write_rows_xlsx(out_xlsx, headers, rows, metadata=[
            ["exported_at", datetime.now().isoformat(timespec="seconds")],
            ["endpoint", "/api/social-posts"],
            ["ranking_formula", "likes + replies*1.5 + reposts*2 + quotes*2 + bookmarks*1.5 + views*0.01"],
            ["date_filter_cn", date_filter or ""],
            ["fetch_meta", json.dumps(fetch_meta, ensure_ascii=False)],
        ])
    if out_json or not out_xlsx:
        write_rows_json(out_json, headers, rows)
    return {"rows": len(ranked), "out_xlsx": out_xlsx, "out_json": out_json, "fetch_meta": fetch_meta}


def main() -> None:
    parser = argparse.ArgumentParser(description="Query Intelligence Engine API")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="API base URL")
    parser.add_argument("--timeout", type=int, default=30, help="request timeout seconds")
    parser.add_argument("--list-aliases", action="store_true", help="list endpoint aliases and exit")
    sub = parser.add_subparsers(dest="cmd")

    get_p = sub.add_parser("get", help="GET an endpoint alias or raw path")
    get_p.add_argument("endpoint", help="endpoint alias or /raw/path")
    get_p.add_argument("--param", action="append", default=[], help="query parameter as key=value; repeatable")
    get_p.add_argument("--out", help="write JSON to file")
    get_p.add_argument("--compact", action="store_true", help="print compact JSON")

    snap_p = sub.add_parser("poster-snapshot", help="collect common poster/briefing data")
    snap_p.add_argument("--date", default=date.today().isoformat(), help="target date YYYY-MM-DD")
    snap_p.add_argument("--days", type=int, default=7, help="lookback window in days")
    snap_p.add_argument("--out", help="write JSON to file")
    snap_p.add_argument("--compact", action="store_true", help="print compact JSON")

    export_p = sub.add_parser("export-papers-xlsx", help="export /api/papers results to an Excel workbook")
    export_p.add_argument("--out", required=True, help="output .xlsx path")
    export_p.add_argument("--venue", help="venue filter, e.g. AAAI")
    export_p.add_argument("--venue-year", help="venue year filter, e.g. 2026")
    export_p.add_argument("--q", help="keyword query")
    export_p.add_argument("--source-name", help="source name filter")
    export_p.add_argument("--source-id", help="source id filter")
    export_p.add_argument("--affiliation", help="affiliation text filter")
    export_p.add_argument("--date-from", help="publication start date")
    export_p.add_argument("--date-to", help="publication end date")
    export_p.add_argument("--has-abstract", help="true/false")
    export_p.add_argument("--sort-by", default="publication_date", help="sort field")
    export_p.add_argument("--order", default="desc", help="sort order")
    export_p.add_argument("--compact", action="store_true", help="print compact JSON summary")

    generic_xlsx_p = sub.add_parser("export-xlsx", help="export any GET endpoint collection to Excel")
    generic_xlsx_p.add_argument("endpoint", help="endpoint alias or /raw/path")
    generic_xlsx_p.add_argument("--param", action="append", default=[], help="query parameter as key=value; repeatable")
    generic_xlsx_p.add_argument("--out", required=True, help="output .xlsx path")
    generic_xlsx_p.add_argument("--items-path", default="items", help="dotted path to list in response; default: items")
    generic_xlsx_p.add_argument("--pagination", choices=["none", "page", "offset"], default="none", help="pagination strategy")
    generic_xlsx_p.add_argument("--page-size", type=int, default=100, help="page size/limit for paginated fetches")
    generic_xlsx_p.add_argument("--max-pages", type=int, help="maximum pages to fetch")
    generic_xlsx_p.add_argument("--fields", help="comma list of paths or label=path, e.g. title,source=source.name")
    generic_xlsx_p.add_argument("--where", action="append", default=[], help="local filter like field~text, field=value, field>10; repeatable")
    generic_xlsx_p.add_argument("--sort", help="local sort field path")
    generic_xlsx_p.add_argument("--sort-desc", action="store_true", help="sort descending")
    generic_xlsx_p.add_argument("--limit", type=int, help="limit exported rows after local filtering/sorting")
    generic_xlsx_p.add_argument("--compact", action="store_true", help="print compact JSON summary")

    chart_p = sub.add_parser("chart", help="fetch endpoint data and export a simple chart image")
    chart_p.add_argument("endpoint", help="endpoint alias or /raw/path")
    chart_p.add_argument("--param", action="append", default=[], help="query parameter as key=value; repeatable")
    chart_p.add_argument("--out", required=True, help="output image path, e.g. chart.png")
    chart_p.add_argument("--x", required=True, help="x/group field path")
    chart_p.add_argument("--y", help="numeric y field path; omit for counts")
    chart_p.add_argument("--chart-type", choices=["bar", "line", "pie"], default="bar")
    chart_p.add_argument("--aggregate", choices=["count", "sum"], default="count")
    chart_p.add_argument("--top", type=int, default=20, help="top groups to keep")
    chart_p.add_argument("--items-path", default="items", help="dotted path to list in response; default: items")
    chart_p.add_argument("--pagination", choices=["none", "page", "offset"], default="none", help="pagination strategy")
    chart_p.add_argument("--page-size", type=int, default=100, help="page size/limit for paginated fetches")
    chart_p.add_argument("--max-pages", type=int, help="maximum pages to fetch")
    chart_p.add_argument("--compact", action="store_true", help="print compact JSON summary")

    rank_p = sub.add_parser("rank-social", help="rank social posts by local heat score and optionally export")
    rank_p.add_argument("--param", action="append", default=[], help="extra /api/social-posts parameter as key=value; repeatable")
    rank_p.add_argument("--date", help="filter by Beijing date YYYY-MM-DD")
    rank_p.add_argument("--limit", type=int, default=50, help="ranked rows to return/export")
    rank_p.add_argument("--max-pages", type=int, default=5, help="pages of social-posts to fetch before ranking")
    rank_p.add_argument("--out-xlsx", help="optional Excel output path")
    rank_p.add_argument("--out-json", help="optional JSON output path")
    rank_p.add_argument("--compact", action="store_true", help="print compact JSON summary")

    args = parser.parse_args()
    if args.list_aliases:
        for name, path in sorted(ALIASES.items()):
            print(f"{name}\t{path}")
        return
    if args.cmd == "get":
        data = api_get(args.base_url, args.endpoint, parse_params(args.param), timeout=args.timeout)
        write_json(data, args.out, pretty=not args.compact)
    elif args.cmd == "poster-snapshot":
        data = poster_snapshot(args.base_url, args.date, args.days, timeout=args.timeout)
        write_json(data, args.out, pretty=not args.compact)
    elif args.cmd == "export-papers-xlsx":
        data = export_papers_xlsx(
            args.base_url,
            args.out,
            args.timeout,
            venue=args.venue,
            venue_year=args.venue_year,
            q=args.q,
            source_name=args.source_name,
            source_id=args.source_id,
            affiliation=args.affiliation,
            date_from=args.date_from,
            date_to=args.date_to,
            has_abstract=args.has_abstract,
            sort_by=args.sort_by,
            order=args.order,
        )
        write_json(data, None, pretty=not args.compact)
    elif args.cmd == "export-xlsx":
        data = generic_export_xlsx(
            args.base_url,
            args.endpoint,
            parse_params(args.param),
            args.out,
            args.timeout,
            items_path=args.items_path,
            pagination=args.pagination,
            page_size=args.page_size,
            max_pages=args.max_pages,
            fields=args.fields,
            where=args.where,
            sort=args.sort,
            sort_desc=args.sort_desc,
            limit=args.limit,
        )
        write_json(data, None, pretty=not args.compact)
    elif args.cmd == "chart":
        data = generic_chart(
            args.base_url,
            args.endpoint,
            parse_params(args.param),
            args.out,
            args.timeout,
            x_field=args.x,
            y_field=args.y,
            chart_type=args.chart_type,
            aggregate=args.aggregate,
            top=args.top,
            items_path=args.items_path,
            pagination=args.pagination,
            page_size=args.page_size,
            max_pages=args.max_pages,
        )
        write_json(data, None, pretty=not args.compact)
    elif args.cmd == "rank-social":
        data = rank_social_posts(
            args.base_url,
            parse_params(args.param),
            args.timeout,
            date_filter=args.date,
            limit=args.limit,
            out_xlsx=args.out_xlsx,
            out_json=args.out_json,
            max_pages=args.max_pages,
        )
        if args.out_xlsx:
            write_json(data, None, pretty=not args.compact)
    else:
        parser.print_help(sys.stderr)
        raise SystemExit(2)


if __name__ == "__main__":
    main()
